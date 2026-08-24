import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    @staticmethod
    def generate_excel_reports(results_data: list, output_dir: str):
        excel_dir = os.path.join(output_dir, "Excel")
        os.makedirs(excel_dir, exist_ok=True)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        accent_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        font_pass = Font(name="Calibri", size=11, bold=True, color="15803D")
        font_fail = Font(name="Calibri", size=11, bold=True, color="B91C1C")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # ---------------------------------------------------------
        # 1. Automation_Test_Report.xlsx (7 Sheets)
        # ---------------------------------------------------------
        wb_master = openpyxl.Workbook()
        
        # Sheet 1: Executed Test Cases
        ws_executed = wb_master.active
        ws_executed.title = "Executed Test Cases"
        headers_1 = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time"]
        for c_idx, h in enumerate(headers_1, 1):
            cell = ws_executed.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for r_idx, item in enumerate(results_data, 2):
            ws_executed.cell(row=r_idx, column=1, value=item["id"]).alignment = Alignment(horizontal="center")
            ws_executed.cell(row=r_idx, column=2, value=item["module"])
            ws_executed.cell(row=r_idx, column=3, value=item["name"])
            ws_executed.cell(row=r_idx, column=4, value=item["priority"]).alignment = Alignment(horizontal="center")
            
            st_cell = ws_executed.cell(row=r_idx, column=5, value=item["status"])
            st_cell.alignment = Alignment(horizontal="center")
            st_cell.font = font_pass if item["status"] == "PASSED" else font_fail
            st_cell.fill = pass_fill if item["status"] == "PASSED" else fail_fill

            ws_executed.cell(row=r_idx, column=6, value=f"{item['duration']:.2f}s").alignment = Alignment(horizontal="center")

            for c_idx in range(1, 7):
                ws_executed.cell(row=r_idx, column=c_idx).border = thin_border

        # Sheet 2: Passed Tests
        ws_passed = wb_master.create_sheet(title="Passed Tests")
        for c_idx, h in enumerate(headers_1, 1):
            cell = ws_passed.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = thin_border
        p_row = 2
        for item in results_data:
            if item["status"] == "PASSED":
                ws_passed.cell(row=p_row, column=1, value=item["id"])
                ws_passed.cell(row=p_row, column=2, value=item["module"])
                ws_passed.cell(row=p_row, column=3, value=item["name"])
                ws_passed.cell(row=p_row, column=4, value=item["priority"])
                ws_passed.cell(row=p_row, column=5, value="PASSED").font = font_pass
                ws_passed.cell(row=p_row, column=6, value=f"{item['duration']:.2f}s")
                p_row += 1

        # Sheet 3: Failed Tests
        ws_failed = wb_master.create_sheet(title="Failed Tests")
        headers_f = ["Test ID", "Module", "Test Name", "Priority", "Failure Reason", "Screenshot Path"]
        for c_idx, h in enumerate(headers_f, 1):
            cell = ws_failed.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = thin_border
        f_row = 2
        for item in results_data:
            if item["status"] == "FAILED":
                ws_failed.cell(row=f_row, column=1, value=item["id"])
                ws_failed.cell(row=f_row, column=2, value=item["module"])
                ws_failed.cell(row=f_row, column=3, value=item["name"])
                ws_failed.cell(row=f_row, column=4, value=item["priority"])
                ws_failed.cell(row=f_row, column=5, value=item.get("reason", "Assertion Failure"))
                ws_failed.cell(row=f_row, column=6, value=item.get("screenshot", ""))
                f_row += 1

        # Sheet 4: Skipped Tests
        ws_skipped = wb_master.create_sheet(title="Skipped Tests")
        for c_idx, h in enumerate(headers_1, 1):
            cell = ws_skipped.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = thin_border

        # Sheet 5: Execution Metrics
        ws_metrics = wb_master.create_sheet(title="Execution Metrics")
        ws_metrics.cell(row=1, column=1, value="Metric Name").font = font_bold
        ws_metrics.cell(row=1, column=2, value="Value").font = font_bold
        total_tc = len(results_data)
        passed_tc = sum(1 for x in results_data if x["status"] == "PASSED")
        failed_tc = sum(1 for x in results_data if x["status"] == "FAILED")
        pass_rate = (passed_tc / total_tc * 100) if total_tc > 0 else 0

        m_data = [
            ("Total Test Cases", total_tc),
            ("Passed Test Cases", passed_tc),
            ("Failed Test Cases", failed_tc),
            ("Skipped Test Cases", 0),
            ("Pass Percentage", f"{pass_rate:.1f}%")
        ]
        for r, (k, v) in enumerate(m_data, 2):
            ws_metrics.cell(row=r, column=1, value=k)
            ws_metrics.cell(row=r, column=2, value=v)

        # Sheet 6: Defect Summary
        ws_defects = wb_master.create_sheet(title="Defect Summary")
        ws_defects.cell(row=1, column=1, value="Defect ID").font = font_bold
        ws_defects.cell(row=1, column=2, value="Description").font = font_bold

        # Sheet 7: Pass Rate Summary
        ws_pass_rate = wb_master.create_sheet(title="Pass Rate Summary")
        ws_pass_rate.cell(row=1, column=1, value="Module").font = font_bold
        ws_pass_rate.cell(row=1, column=2, value="Pass Rate").font = font_bold

        # Auto-fit columns for all sheets in master
        for sheet in wb_master.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(c.value or '')) for c in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        master_path = os.path.join(excel_dir, "Automation_Test_Report.xlsx")
        wb_master.save(master_path)

        # ---------------------------------------------------------
        # 2. Passed_Test_Cases.xlsx
        # ---------------------------------------------------------
        wb_p = openpyxl.Workbook()
        ws_p = wb_p.active
        ws_p.title = "Passed Test Cases"
        for c_idx, h in enumerate(headers_1, 1):
            cell = ws_p.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
        p_row = 2
        for item in results_data:
            if item["status"] == "PASSED":
                ws_p.cell(row=p_row, column=1, value=item["id"])
                ws_p.cell(row=p_row, column=2, value=item["module"])
                ws_p.cell(row=p_row, column=3, value=item["name"])
                ws_p.cell(row=p_row, column=4, value=item["priority"])
                ws_p.cell(row=p_row, column=5, value="PASSED").font = font_pass
                ws_p.cell(row=p_row, column=6, value=f"{item['duration']:.2f}s")
                p_row += 1
        wb_p.save(os.path.join(excel_dir, "Passed_Test_Cases.xlsx"))

        # ---------------------------------------------------------
        # 3. Failed_Test_Cases.xlsx
        # ---------------------------------------------------------
        wb_f = openpyxl.Workbook()
        ws_f = wb_f.active
        ws_f.title = "Failed Test Cases"
        for c_idx, h in enumerate(headers_f, 1):
            cell = ws_f.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
        f_row = 2
        for item in results_data:
            if item["status"] == "FAILED":
                ws_f.cell(row=f_row, column=1, value=item["id"])
                ws_f.cell(row=f_row, column=2, value=item["module"])
                ws_f.cell(row=f_row, column=3, value=item["name"])
                ws_f.cell(row=f_row, column=4, value=item["priority"])
                ws_f.cell(row=f_row, column=5, value=item.get("reason", "Assertion Failure"))
                ws_f.cell(row=f_row, column=6, value=item.get("screenshot", ""))
                f_row += 1
        wb_f.save(os.path.join(excel_dir, "Failed_Test_Cases.xlsx"))

        # ---------------------------------------------------------
        # 4. Execution_Summary.xlsx
        # ---------------------------------------------------------
        wb_s = openpyxl.Workbook()
        ws_s = wb_s.active
        ws_s.title = "Execution Summary"
        ws_s.cell(row=1, column=1, value="Metric").font = font_bold
        ws_s.cell(row=1, column=2, value="Count").font = font_bold
        for r, (k, v) in enumerate(m_data, 2):
            ws_s.cell(row=r, column=1, value=k)
            ws_s.cell(row=r, column=2, value=v)
        wb_s.save(os.path.join(excel_dir, "Execution_Summary.xlsx"))

        print(f"Excel Reports generated successfully in: {excel_dir}")
